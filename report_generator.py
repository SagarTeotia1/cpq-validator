from __future__ import annotations

import io
from datetime import datetime
from typing import Optional

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

from validator import ValidationResult


def generate_report(validation_result: ValidationResult) -> bytes:
    """Generate a simple summary PDF for validation results (Phase 4).

    Later phases can add annotated original PDF overlays.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, title="CPQ Quote Validation Report")
    styles = getSampleStyleSheet()
    elements = []

    # Title
    title_text = f"VALIDATION SUMMARY — {validation_result.overall_status}"
    elements.append(Paragraph(title_text, styles["Title"]))
    elements.append(Spacer(1, 8 * mm))

    # Meta
    ts = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
    meta_lines = [
        f"Validation timestamp: {ts}",
        f"Transaction ID: {validation_result.transaction_id or '-'}",
        f"PDF Filename: {validation_result.pdf_filename or '-'}",
        f"Total fields checked: {validation_result.total_checked}",
        f"Matches: {validation_result.matches}",
        f"Mismatches: {validation_result.mismatches}",
    ]
    for line in meta_lines:
        elements.append(Paragraph(line, styles["Normal"]))
    elements.append(Spacer(1, 10 * mm))

    # Discrepancy table
    data = [["Field Name", "Section", "Expected (API)", "Found (PDF)", "Match"]]
    for d in validation_result.details:
        data.append([
            d.field_name,
            d.section,
            str(d.expected),
            str(d.found),
            "✓" if d.match else "✗",
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("FONT", (0, 0), (-1, -1), "Helvetica", 9),
                ("ALIGN", (-1, 1), (-1, -1), "CENTER"),
            ]
        )
    )
    elements.append(table)

    doc.build(elements)
    return buffer.getvalue()


def generate_xlsx(validation_result: ValidationResult) -> bytes:
    """Generate an .xlsx workbook summarizing validation results.

    Sheet1: Summary
    Sheet2: Field Results
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    ws["A1"] = f"Validation {validation_result.overall_status}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Validation timestamp"
    ws["B3"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
    ws["A4"] = "Transaction ID"
    ws["B4"] = validation_result.transaction_id or "-"
    ws["A5"] = "PDF Filename"
    ws["B5"] = validation_result.pdf_filename or "-"
    ws["A6"] = "Total fields checked"
    ws["B6"] = validation_result.total_checked
    ws["A7"] = "Matches"
    ws["B7"] = validation_result.matches
    ws["A8"] = "Mismatches"
    ws["B8"] = validation_result.mismatches

    # Details sheet
    wd = wb.create_sheet("Details")
    headers = ["Field Name", "Section", "Expected (API)", "Found (PDF)", "Match"]
    wd.append(headers)
    bold = Font(bold=True)
    for col in range(1, len(headers) + 1):
        wd.cell(row=1, column=col).font = bold
    red_fill = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")
    green_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")

    for d in validation_result.details:
        row = [d.field_name, d.section, str(d.expected), str(d.found), "✓" if d.match else "✗"]
        wd.append(row)
        r_idx = wd.max_row
        if d.match:
            for c in range(1, 6):
                wd.cell(row=r_idx, column=c).fill = green_fill
        else:
            for c in range(1, 6):
                wd.cell(row=r_idx, column=c).fill = red_fill

    # Autosize columns
    for sheet in [ws, wd]:
        for col in sheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            sheet.column_dimensions[col_letter].width = min(max_len + 2, 60)

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


