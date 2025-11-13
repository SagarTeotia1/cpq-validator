"""Convert API JSON response to Excel format for comparison"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any, Dict, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def extract_line_items(api_data: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Extract line items from API response."""
    line_items = []
    
    # Try different structures
    lines_container = api_data.get("transactionLine") or {}
    if isinstance(lines_container, dict) and "items" in lines_container:
        items = lines_container.get("items", [])
    elif isinstance(api_data.get("items"), list):
        items = api_data.get("items", [])
    else:
        items = []
    
    for item in items:
        # Extract part number
        part_number = (
            item.get("_part_number") or 
            item.get("_part_display_number") or 
            item.get("_line_display_name") or
            item.get("partNumber") or
            ""
        )
        
        # Extract description
        description = (
            item.get("_line_description") or
            item.get("description") or
            item.get("_product_description") or
            ""
        )
        
        # Extract quantity
        quantity = (
            item.get("_price_quantity") or
            item.get("_line_bom_item_quantity") or
            item.get("quantity") or
            0
        )
        
        # Extract unit list price
        unit_list_price = None
        for key in ["_price_item_price_each", "_price_unit_price_each", "_price_list_price_each", "unitListPrice"]:
            val = item.get(key)
            if isinstance(val, dict) and val.get("value") is not None:
                unit_list_price = val.get("value")
                break
            elif isinstance(val, (int, float)) and val is not None:
                unit_list_price = val
                break
        
        # Extract unit net price
        unit_net_price = None
        net_price_obj = item.get("netPrice_l") or item.get("unitNetPrice")
        if isinstance(net_price_obj, dict) and net_price_obj.get("value") is not None:
            unit_net_price = net_price_obj.get("value")
        elif isinstance(net_price_obj, (int, float)) and net_price_obj is not None:
            unit_net_price = net_price_obj
        
        # Extract extended net price
        extended_net_price = None
        for key in ["netAmount_l", "netAmountRollup_l", "netPriceRollup_l", "extendedNetPrice"]:
            val = item.get(key)
            if isinstance(val, dict) and val.get("value") is not None:
                extended_net_price = val.get("value")
                break
            elif isinstance(val, (int, float)) and val is not None:
                extended_net_price = val
                break
        
        # Extract extended list price
        extended_list_price = None
        for key in ["_price_extended_price", "extendedListPrice", "listAmount_l"]:
            val = item.get(key)
            if isinstance(val, dict) and val.get("value") is not None:
                extended_list_price = val.get("value")
                break
            elif isinstance(val, (int, float)) and val is not None:
                extended_list_price = val
                break
        
        # Calculate if missing
        if extended_list_price is None and unit_list_price is not None and quantity:
            extended_list_price = float(unit_list_price) * float(quantity)
        
        if extended_net_price is None and unit_net_price is not None and quantity:
            extended_net_price = float(unit_net_price) * float(quantity)
        
        # Extract discount percent
        discount_percent = item.get("discountPercent_l") or item.get("discountPercent")
        if isinstance(discount_percent, dict):
            discount_percent = discount_percent.get("value")
        
        # Calculate discount if missing
        if discount_percent is None and unit_list_price and unit_net_price:
            try:
                discount_percent = ((float(unit_list_price) - float(unit_net_price)) / float(unit_list_price)) * 100
            except (ValueError, TypeError, ZeroDivisionError):
                discount_percent = None
        
        line_items.append({
            "Part Number": part_number,
            "Description": description,
            "Quantity": quantity,
            "Unit List Price": unit_list_price,
            "Unit Net Price": unit_net_price,
            "Extended List Price": extended_list_price,
            "Extended Net Price": extended_net_price,
            "Discount %": discount_percent,
        })
    
    return line_items


def json_to_excel(json_file: str, output_file: str):
    """Convert JSON API response to Excel format."""
    # Load JSON
    with open(json_file, "r", encoding="utf-8") as f:
        api_data = json.load(f)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "API Data"
    
    # Header information
    row = 1
    
    # Quote Header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    label_font = Font(bold=True, size=11)
    data_font = Font(size=11)
    
    ws.merge_cells(f"A{row}:B{row}")
    cell = ws[f"A{row}"]
    cell.value = "Quote Information"
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 2
    
    # Quote fields
    fields = [
        ("Quote Number", api_data.get("quoteNumber_t_c") or api_data.get("_document_number")),
        ("Transaction ID", api_data.get("transactionID_t") or api_data.get("quoteTransactionID_t_c") or api_data.get("bs_id")),
        ("Quote Name", api_data.get("quoteNameTextArea_t_c") or api_data.get("transactionName_t")),
        ("Status", api_data.get("quoteStatus_t_c") or api_data.get("status_t")),
        ("Quote Date", api_data.get("createdDate_t")),
        ("Valid Until", api_data.get("expiresOnDate_t_c")),
        ("Currency", api_data.get("currency_t")),
        ("Price List", api_data.get("priceList_t_c")),
        ("Incoterm", api_data.get("incoterm_t_c")),
        ("Payment Terms", api_data.get("paymentTerms_t_c")),
        ("Order Type", api_data.get("orderType_t_c")),
    ]
    
    for label, value in fields:
        ws[f"A{row}"] = label + ":"
        ws[f"A{row}"].font = label_font
        ws[f"B{row}"] = str(value) if value is not None else ""
        ws[f"B{row}"].font = data_font
        row += 1
    
    row += 1
    
    # Totals Section
    ws.merge_cells(f"A{row}:B{row}")
    cell = ws[f"A{row}"]
    cell.value = "Pricing Summary"
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 2
    
    # Extract totals
    list_price = api_data.get("quoteListPrice_t_c") or api_data.get("totalOneTimeListAmount_t")
    if isinstance(list_price, dict):
        list_price = list_price.get("value")
    
    net_price = api_data.get("quoteNetPrice_t_c") or api_data.get("totalOneTimeNetAmount_t") or api_data.get("_transaction_total")
    if isinstance(net_price, dict):
        net_price = net_price.get("value")
    
    discount = api_data.get("quoteCurrentDiscount_t_c") or api_data.get("transactionTotalDiscountPercent_t")
    if isinstance(discount, dict):
        discount = discount.get("value")
    
    # Calculate discount if missing
    if discount is None and list_price and net_price:
        try:
            discount = ((float(list_price) - float(net_price)) / float(list_price)) * 100
        except (ValueError, TypeError, ZeroDivisionError):
            discount = None
    
    totals = [
        ("List Grand Total", list_price),
        ("Total Discount (%)", discount),
        ("Net Grand Total", net_price),
    ]
    
    total_font = Font(bold=True, size=11)
    for label, value in totals:
        ws[f"A{row}"] = label + ":"
        ws[f"A{row}"].font = total_font
        if isinstance(value, (int, float)):
            ws[f"B{row}"] = float(value)
            ws[f"B{row}"].number_format = "#,##0.00" if label.endswith("%") else "$#,##0.00"
        else:
            ws[f"B{row}"] = str(value) if value is not None else ""
        ws[f"B{row}"].font = total_font
        row += 1
    
    row += 1
    
    # Line Items
    line_items = extract_line_items(api_data)
    
    if line_items:
        ws.merge_cells(f"A{row}:H{row}")
        cell = ws[f"A{row}"]
        cell.value = "Line Items"
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1
        
        # Headers
        headers = ["Part Number", "Description", "Quantity", "Unit List Price", 
                   "Unit Net Price", "Extended List Price", "Extended Net Price", "Discount %"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = label_font
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        row += 1
        
        # Data rows
        for item in line_items:
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                value = item.get(header, "")
                
                if header in ["Unit List Price", "Unit Net Price", "Extended List Price", "Extended Net Price"]:
                    if value is not None:
                        cell.value = float(value)
                        cell.number_format = "$#,##0.00"
                    else:
                        cell.value = ""
                elif header == "Discount %":
                    if value is not None:
                        cell.value = float(value)
                        cell.number_format = "0.00"
                    else:
                        cell.value = ""
                elif header == "Quantity":
                    cell.value = int(value) if value else 0
                else:
                    cell.value = str(value) if value else ""
                
                cell.font = data_font
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            row += 1
        
        # Calculate totals from line items
        row += 1
        ws.merge_cells(f"A{row}:E{row}")
        cell = ws[f"A{row}"]
        cell.value = "Calculated Totals from Line Items:"
        cell.font = total_font
        
        row += 1
        calculated_list_total = sum(float(item.get("Extended List Price", 0) or 0) for item in line_items)
        calculated_net_total = sum(float(item.get("Extended Net Price", 0) or 0) for item in line_items)
        calculated_discount = ((calculated_list_total - calculated_net_total) / calculated_list_total * 100) if calculated_list_total > 0 else 0
        
        calc_totals = [
            ("Calculated List Total", calculated_list_total),
            ("Calculated Net Total", calculated_net_total),
            ("Calculated Discount %", calculated_discount),
        ]
        
        for label, value in calc_totals:
            ws[f"A{row}"] = label + ":"
            ws[f"A{row}"].font = total_font
            ws[f"B{row}"] = float(value)
            if label.endswith("%"):
                ws[f"B{row}"].number_format = "0.00"
            else:
                ws[f"B{row}"].number_format = "$#,##0.00"
            ws[f"B{row}"].font = total_font
            row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 12
    
    # Save
    wb.save(output_file)
    print(f"[OK] Excel file created: {output_file}")


def main():
    parser = argparse.ArgumentParser(description="Convert JSON API response to Excel")
    parser.add_argument("--json", default="response.json", help="Input JSON file (default: response.json)")
    parser.add_argument("--output", help="Output Excel file (default: api_data.xlsx)")
    args = parser.parse_args()
    
    json_file = Path(args.json)
    if not json_file.exists():
        print(f"ERROR: JSON file not found: {json_file}")
        sys.exit(1)
    
    output_file = args.output or "api_data.xlsx"
    
    print("="*60)
    print("JSON to Excel Converter")
    print("="*60)
    print(f"Input:  {json_file}")
    print(f"Output: {output_file}")
    print()
    
    try:
        json_to_excel(str(json_file), output_file)
        print("\n[OK] Conversion complete!")
    except Exception as e:
        print(f"\n[ERROR] Conversion failed: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()

